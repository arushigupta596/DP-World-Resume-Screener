from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SHORTLIST_FILL = PatternFill("solid", fgColor="E1F5EE")
REVIEW_FILL = PatternFill("solid", fgColor="FAEEDA")
REJECT_FILL = PatternFill("solid", fgColor="FCEBEB")

HEADERS = [
    "Rank",
    "Name",
    "File",
    "Total Score",
    "Recommendation",
    "C1",
    "C2",
    "C3",
    "C4",
    "C5",
    "Risk Flags",
    "Bonus Tools",
    "AI Summary",
]


def _fill_for(recommendation: str) -> PatternFill | None:
    if recommendation == "Shortlist":
        return SHORTLIST_FILL
    if recommendation == "Review":
        return REVIEW_FILL
    if recommendation == "Reject":
        return REJECT_FILL
    return None


def generate_excel(role: dict, candidates_with_scores: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Scores"

    header_font = Font(bold=True)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"

    ranked = sorted(
        candidates_with_scores,
        key=lambda c: (c.get("score") or {}).get("total_score") or -1,
        reverse=True,
    )

    for rank, candidate in enumerate(ranked, start=1):
        score = candidate.get("score") or {}
        criteria = score.get("criteria_scores") or {}
        row = [
            rank,
            candidate.get("name") or "",
            candidate.get("file_name") or "",
            score.get("total_score") if score.get("total_score") is not None else "",
            score.get("recommendation") or candidate.get("status") or "",
            (criteria.get("C1") or {}).get("score", ""),
            (criteria.get("C2") or {}).get("score", ""),
            (criteria.get("C3") or {}).get("score", ""),
            (criteria.get("C4") or {}).get("score", ""),
            (criteria.get("C5") or {}).get("score", ""),
            ", ".join(score.get("risk_flags") or []),
            ", ".join(score.get("bonus_tools") or []),
            score.get("ai_summary") or "",
        ]
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=rank + 1, column=col_idx, value=value)

        fill = _fill_for(score.get("recommendation") or "")
        if fill:
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=rank + 1, column=col_idx).fill = fill

    widths = [6, 22, 28, 12, 14, 6, 6, 6, 6, 6, 36, 24, 60]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
